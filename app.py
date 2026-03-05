"""
Simulateur Jeanbrun – Streamlit App
Architecture : LibreOffice headless recalcule le fichier Excel,
puis on lit les valeurs calculées avec openpyxl.
"""

import streamlit as st
import openpyxl
import shutil
import subprocess
import os
import tempfile
from pathlib import Path
import pandas as pd
import io

# ─────────────────────────────────────────────
# CONFIG PAGE
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Simulateur Jeanbrun",
    page_icon="🏠",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────
st.markdown("""
<style>
/* Palette couleurs */
:root {
    --dark-blue:  #13415B;
    --med-blue:   #3761AD;
    --teal:       #009FA3;
    --orange:     #EA653D;
    --light-blue: #EEF2FB;
    --light-teal: #E5F6F6;
    --light-bg:   #F4F6F9;
}

/* En-tête */
.header-band {
    background: linear-gradient(135deg, var(--dark-blue) 0%, var(--med-blue) 100%);
    color: white;
    padding: 1.2rem 2rem;
    border-radius: 10px;
    margin-bottom: 1.5rem;
}
.header-band h1 { margin: 0; font-size: 1.6rem; }
.header-band p  { margin: 0.3rem 0 0; opacity: .8; font-size: .9rem; }

/* Cards KPI */
.kpi-row { display: flex; gap: 1rem; flex-wrap: wrap; margin: 1rem 0; }
.kpi-card {
    flex: 1; min-width: 160px;
    background: var(--light-blue);
    border-left: 4px solid var(--med-blue);
    border-radius: 8px;
    padding: 0.9rem 1.1rem;
}
.kpi-card.teal   { background: var(--light-teal); border-color: var(--teal); }
.kpi-card.orange { background: #FFF3EE; border-color: var(--orange); }
.kpi-card.dark   { background: #E8EDF4; border-color: var(--dark-blue); }
.kpi-label { font-size: .75rem; color: #555; text-transform: uppercase; letter-spacing: .05em; }
.kpi-value { font-size: 1.4rem; font-weight: 700; color: var(--dark-blue); margin-top: .2rem; }
.kpi-sub   { font-size: .75rem; color: #777; margin-top: .1rem; }

/* Effort badge */
.effort-badge {
    display: inline-block;
    background: var(--dark-blue);
    color: white;
    padding: 0.5rem 1.4rem;
    border-radius: 30px;
    font-size: 1.1rem;
    font-weight: 700;
    margin: 0.5rem 0;
}

/* Tables */
.stDataFrame { border-radius: 8px; overflow: hidden; }
thead tr th { background: var(--dark-blue) !important; color: white !important; }

/* Section header */
.section-header {
    background: var(--dark-blue);
    color: white;
    padding: .5rem 1rem;
    border-radius: 6px;
    font-weight: 600;
    margin: 1.2rem 0 .6rem;
}

/* Login */
.login-wrap {
    max-width: 420px;
    margin: 5rem auto;
    padding: 2.5rem;
    background: white;
    border-radius: 12px;
    box-shadow: 0 4px 24px rgba(19,65,91,.12);
    text-align: center;
}

/* Sidebar section */
.sidebar-section {
    background: var(--light-bg);
    border-radius: 8px;
    padding: .6rem .8rem;
    margin-bottom: .8rem;
    border-left: 3px solid var(--med-blue);
}

/* Print */
@media print {
    .stSidebar, .stButton, [data-testid="stToolbar"],
    [data-testid="stDecoration"], .block-container > *:first-child { display: none !important; }
    .block-container { padding: 0 !important; max-width: 100% !important; }
    .print-page { page-break-after: always; }
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# AUTHENTIFICATION
# ─────────────────────────────────────────────
def check_password() -> bool:
    if st.session_state.get("auth"):
        return True

    st.markdown("""
    <div class="login-wrap">
        <h2 style="color:#13415B; margin-top:0">🏠 Simulateur<br>Dispositif Jeanbrun</h2>
        <p style="color:#555">Outil réservé aux conseillers</p>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        pwd = st.text_input("Mot de passe", type="password", label_visibility="collapsed",
                            placeholder="Entrez le mot de passe")
        if st.button("🔓 Se connecter", use_container_width=True):
            correct = st.secrets.get("password", "jeanbrun2025")
            if pwd == correct:
                st.session_state.auth = True
                st.rerun()
            else:
                st.error("Mot de passe incorrect")
    return False


if not check_password():
    st.stop()

# ─────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────
TEMPLATE = Path("Simulation_JEANBRUN_V9.xlsx")


# ─────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────
st.markdown("""
<div class="header-band">
  <h1>🏠 Simulateur — Dispositif Jeanbrun</h1>
  <p>Outil de projection fiscale · Réservé aux conseillers · Document non contractuel</p>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# SIDEBAR — FORMULAIRE ENTRÉES
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ✏️ Paramètres conseiller")
    st.caption("Modifiez uniquement ces champs — les calculs sont automatiques")

    st.markdown("### 🏠 Bien immobilier")
    prix_acq = st.number_input("Prix d'acquisition (€)", min_value=50_000, max_value=5_000_000,
                                value=260_000, step=1_000, format="%d")
    frais_acq = st.number_input("Frais de notaire (%)", min_value=0.0, max_value=15.0,
                                 value=3.0, step=0.1, format="%.1f") / 100
    surface = st.number_input("Surface habitable (m²)", min_value=5.0, max_value=500.0,
                               value=40.0, step=0.5, format="%.1f")
    zone = st.selectbox("Zone d'acquisition", ["A bis", "A", "B1", "B2", "C"], index=1)
    rdc = st.selectbox("Rez-de-chaussée ?", ["NON", "OUI"], index=0)
    balcon = st.number_input("Surface balcon (m²)", min_value=0.0, max_value=100.0,
                              value=15.0, step=0.5, format="%.1f")
    terrasse = st.number_input("Surface terrasse (m²)", min_value=0.0, max_value=300.0,
                                value=0.0, step=0.5, format="%.1f")

    st.markdown("### 💳 Financement")
    apport = st.number_input("Apport personnel (€)", min_value=0, max_value=2_000_000,
                              value=15_000, step=500, format="%d")
    taux_int = st.number_input("Taux d'intérêt annuel (%)", min_value=0.0, max_value=10.0,
                                value=3.3, step=0.05, format="%.2f") / 100
    taux_ass = st.number_input("Taux assurance emprunteur (%)", min_value=0.0, max_value=3.0,
                                value=0.35, step=0.01, format="%.2f") / 100
    duree = st.number_input("Durée du financement (ans)", min_value=5, max_value=30,
                             value=25, step=1)
    frais_garan = st.number_input("Frais garantie + dossier (€)", min_value=0, max_value=20_000,
                                   value=4_000, step=100, format="%d")

    st.markdown("### 🏘️ Revenus locatifs")
    type_loyer = st.selectbox("Type de loyer",
                               ["Loyer intermédiaire", "Loyer social", "Loyer très social"])
    loyer_souhaite = st.number_input("Loyer souhaité (€/mois)", min_value=100, max_value=5_000,
                                      value=750, step=10, format="%d")
    index_loyer = st.number_input("Indexation loyers (%/an)", min_value=0.0, max_value=5.0,
                                   value=1.5, step=0.1, format="%.1f") / 100
    charges_pct = st.number_input("Charges + taxe foncière (% loyers)", min_value=0.0,
                                   max_value=60.0, value=30.0, step=1.0, format="%.0f") / 100

    st.markdown("### 👤 Situation fiscale du client")
    type_revenus = st.selectbox("Type de revenus",
                                 ["Salaires (abatt. 10%)", "BNC / BIC / autres"],
                                 index=0)
    revenus = st.number_input("Revenus annuels déclarés (€)", min_value=0, max_value=2_000_000,
                               value=95_000, step=1_000, format="%d")
    rf_autres = st.number_input("Revenus fonciers autres biens (€/an)", min_value=0,
                                 max_value=500_000, value=5_000, step=500, format="%d")
    parts = st.number_input("Nombre de parts fiscales", min_value=1.0, max_value=10.0,
                             value=2.5, step=0.5, format="%.1f")
    nb_declarants = st.number_input("Nombre de déclarants", min_value=1, max_value=2,
                                     value=2, step=1)

    st.divider()
    calc_btn = st.button("🚀 Lancer la simulation", use_container_width=True, type="primary")

# ─────────────────────────────────────────────
# FONCTION : RECALCUL VIA LIBREOFFICE
# ─────────────────────────────────────────────
@st.cache_data(show_spinner=False, ttl=3600)
def run_simulation(
    prix_acq, frais_acq, surface, zone, rdc, balcon, terrasse,
    apport, taux_int, taux_ass, duree, frais_garan,
    type_loyer, loyer_souhaite, index_loyer, charges_pct,
    type_revenus, revenus, rf_autres, parts, nb_declarants,
):
    """Remplit les cellules bleues dans le template Excel et recalcule via LibreOffice."""

    # 1. Copier le template dans /tmp
    tmp_in  = Path(tempfile.mktemp(suffix=".xlsx"))
    tmp_out = Path(tempfile.mktemp(suffix=".xlsx"))
    shutil.copy(TEMPLATE, tmp_in)

    # 2. Remplir les cellules bleues (Hypothèses)
    wb = openpyxl.load_workbook(tmp_in)
    ws = wb["Hypothèses"]

    ws["B4"]  = float(prix_acq)
    ws["B5"]  = float(frais_acq)
    ws["F5"]  = rdc
    ws["F6"]  = float(terrasse)
    ws["F7"]  = float(balcon)
    ws["B8"]  = float(surface)
    ws["B9"]  = zone
    ws["B14"] = float(apport)
    ws["B16"] = float(taux_int)
    ws["B17"] = float(taux_ass)
    ws["B18"] = float(duree)
    ws["B21"] = float(frais_garan)
    ws["B23"] = type_loyer
    ws["F26"] = float(loyer_souhaite)
    ws["B27"] = float(index_loyer)
    ws["B28"] = float(charges_pct)
    ws["B36"] = type_revenus
    ws["B37"] = float(revenus)
    ws["B38"] = float(rf_autres)
    ws["B39"] = float(parts)
    ws["B41"] = float(nb_declarants)

    wb.save(tmp_in)

    # 3. LibreOffice recalcule et reconvertit en xlsx
    out_dir = tmp_in.parent
    result = subprocess.run(
        [
            "libreoffice", "--headless", "--norestore",
            "--calc", "--convert-to", "xlsx",
            "--outdir", str(out_dir),
            str(tmp_in),
        ],
        capture_output=True, text=True, timeout=90,
    )

    # Le fichier output a le même nom que tmp_in mais .xlsx est déjà son extension
    # LibreOffice le sort dans out_dir avec le même nom
    calc_output = out_dir / tmp_in.name  # même fichier, LibreOffice écrase
    if not calc_output.exists():
        # Chercher le fichier produit
        candidates = sorted(out_dir.glob("*.xlsx"), key=lambda f: f.stat().st_mtime)
        if candidates:
            calc_output = candidates[-1]

    # 4. Lire les valeurs calculées
    wb2 = openpyxl.load_workbook(str(calc_output), data_only=True)

    tmp_in.unlink(missing_ok=True)

    return wb2


def fmt_eur(v, decimals=0):
    if v is None: return "—"
    try:
        if decimals == 0:
            return f"{float(v):,.0f} €".replace(",", " ")
        return f"{float(v):,.{decimals}f} €".replace(",", " ")
    except Exception:
        return str(v)


def fmt_pct(v, decimals=1):
    if v is None: return "—"
    try:
        return f"{float(v)*100:.{decimals}f} %"
    except Exception:
        return str(v)


def fmt_num(v, decimals=0):
    if v is None: return "—"
    try:
        return f"{float(v):,.{decimals}f}".replace(",", " ")
    except Exception:
        return str(v)


# ─────────────────────────────────────────────
# ÉTAT SESSION
# ─────────────────────────────────────────────
if "wb_result" not in st.session_state:
    st.session_state.wb_result = None

if calc_btn:
    with st.spinner("⚙️ Calcul en cours — LibreOffice recalcule toutes les formules…"):
        wb_result = run_simulation(
            prix_acq, frais_acq, surface, zone, rdc, balcon, terrasse,
            apport, taux_int, taux_ass, duree, frais_garan,
            type_loyer, loyer_souhaite, index_loyer, charges_pct,
            type_revenus, revenus, rf_autres, parts, nb_declarants,
        )
    st.session_state.wb_result = wb_result
    st.success("✅ Simulation calculée avec succès !")

# ─────────────────────────────────────────────
# AFFICHAGE DES RÉSULTATS
# ─────────────────────────────────────────────
wb = st.session_state.wb_result

if wb is None:
    st.info("👈 Renseignez les paramètres dans la barre latérale puis cliquez sur **Lancer la simulation**.")
    st.stop()

# Accès aux feuilles
hyp  = wb["Hypothèses"]
simp = wb["Synthèse client - simplifiée"]
det  = wb["Synthèse Client - détaillée"]
rev  = wb["Revente"]
mot  = wb["Moteur"]


def v(ws, cell):
    """Lit une valeur calculée, retourne None si absente."""
    try:
        val = ws[cell].value
        return val
    except Exception:
        return None


# ─────────────────────────────────────────────
# ONGLETS
# ─────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📊 Synthèse Visuelle",
    "📋 Synthèse Simplifiée",
    "📈 Synthèse Détaillée",
    "🏦 Revente & Plus-value",
    "⬇️ Télécharger",
])


# ══════════════════════════════════════════════
# ONGLET 1 — SYNTHÈSE VISUELLE
# ══════════════════════════════════════════════
with tab1:
    st.markdown('<div class="print-page">', unsafe_allow_html=True)

    # KPIs principaux
    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        st.markdown(f"""
        <div class="kpi-card">
          <div class="kpi-label">Revenus déclarés</div>
          <div class="kpi-value">{fmt_eur(v(hyp,'B37'))}</div>
          <div class="kpi-sub">{fmt_num(v(hyp,'B39'), 1)} parts fiscales</div>
        </div>""", unsafe_allow_html=True)
    with col2:
        tmi_val = v(det, 'F6')
        tmi_str = f"{int(float(tmi_val)*100)} %" if tmi_val else "—"
        st.markdown(f"""
        <div class="kpi-card teal">
          <div class="kpi-label">Tranche Marginale</div>
          <div class="kpi-value">{tmi_str}</div>
          <div class="kpi-sub">avant opération</div>
        </div>""", unsafe_allow_html=True)
    with col3:
        st.markdown(f"""
        <div class="kpi-card">
          <div class="kpi-label">Prix d'acquisition</div>
          <div class="kpi-value">{fmt_eur(v(hyp,'B4'))}</div>
          <div class="kpi-sub">{v(hyp,'B23') or '—'}</div>
        </div>""", unsafe_allow_html=True)
    with col4:
        st.markdown(f"""
        <div class="kpi-card dark">
          <div class="kpi-label">Loyer mensuel initial</div>
          <div class="kpi-value">{fmt_eur(v(hyp,'B25'))}</div>
          <div class="kpi-sub">Zone {v(hyp,'B9')} · {fmt_num(v(hyp,'F9'), 1)} m² pond.</div>
        </div>""", unsafe_allow_html=True)
    with col5:
        eco_an1 = v(det, 'F7')
        st.markdown(f"""
        <div class="kpi-card orange">
          <div class="kpi-label">Économie fiscale an 1</div>
          <div class="kpi-value">{fmt_eur(eco_an1)}</div>
          <div class="kpi-sub">déficit foncier + Jeanbrun</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("---")

    # ── Compte en T — 3 horizons ──
    def compte_en_t(label, horizon, entrees, sorties, effort, capital, gain_total, color_class=""):
        ef = float(effort) if effort else 0
        effort_str = fmt_eur(abs(ef))
        effort_color = "#EA653D" if ef < 0 else "#009FA3"
        return f"""
        <div style="background:{'#EEF2FB' if not color_class else '#E5F6F6' if 'teal' in color_class else '#FFF3EE'};
                    border-radius:10px; padding:1.2rem; flex:1; min-width:280px;">
          <div style="font-weight:700; color:#13415B; font-size:1rem; margin-bottom:.8rem;">
            {label} — <span style="color:#3761AD;">{horizon}</span>
          </div>
          <table style="width:100%; border-collapse:collapse; font-size:.88rem;">
            <tr>
              <td style="color:#009FA3;font-weight:600;padding:.3rem;">✚ CE QUI RENTRE</td>
              <td style="color:#EA653D;font-weight:600;padding:.3rem;">— CE QUI SORT</td>
            </tr>
            <tr>
              <td style="padding:.2rem .3rem;">Loyers moy. : <b>{fmt_eur(entrees[0])}/mois</b></td>
              <td style="padding:.2rem .3rem;">Crédit : <b>{fmt_eur(sorties[0])}/mois</b></td>
            </tr>
            <tr>
              <td style="padding:.2rem .3rem;">Gain fiscal moy. : <b>{fmt_eur(entrees[1])}/mois</b></td>
              <td style="padding:.2rem .3rem;">Charges : <b>{fmt_eur(sorties[1])}/mois</b></td>
            </tr>
            <tr style="border-top:1px solid #ccc;font-weight:600;">
              <td style="padding:.3rem;">Total : <b>{fmt_eur(entrees[2])}/mois</b></td>
              <td style="padding:.3rem;">Total : <b>{fmt_eur(sorties[2])}/mois</b></td>
            </tr>
          </table>
          <div style="margin-top:.8rem; text-align:center;">
            <span style="font-size:.8rem; color:#555;">Effort d'investissement mensuel moyen</span><br>
            <span style="font-size:1.3rem; font-weight:700; color:{effort_color};">{effort_str}/mois</span>
          </div>
          <div style="margin-top:.6rem; background:white; border-radius:6px; padding:.6rem; font-size:.82rem;">
            <b>Capital net constitué :</b> {fmt_eur(capital)}<br>
            <b>Gain fiscal total :</b> {fmt_eur(gain_total)}
          </div>
        </div>
        """

    st.markdown("### 📊 Compte en T — Moyennes mensuelles par horizon")

    c9, c15, c25 = st.columns(3)
    with c9:
        st.markdown(compte_en_t(
            "🔹 Fin d'engagement", "9 ans",
            [v(simp,'D14'), v(simp,'D15'), v(simp,'D16')],
            [v(simp,'I14'), v(simp,'I15'), v(simp,'I16')],
            v(simp,'G17'), v(simp,'M14'), v(simp,'M15'),
        ), unsafe_allow_html=True)

    with c15:
        st.markdown(compte_en_t(
            "🔸 Horizon de référence", "15 ans",
            [v(simp,'D22'), v(simp,'D23'), v(simp,'D24')],
            [v(simp,'I22'), v(simp,'I23'), v(simp,'I24')],
            v(simp,'G25'), v(simp,'M22'), v(simp,'M23'), "teal",
        ), unsafe_allow_html=True)

    with c25:
        st.markdown(compte_en_t(
            "⭐ Financement soldé", "25 ans",
            [v(simp,'D30'), v(simp,'D31'), v(simp,'D32')],
            [v(simp,'I30'), v(simp,'I31'), v(simp,'I32')],
            v(simp,'G33'), v(simp,'M30'), v(simp,'M31'), "orange",
        ), unsafe_allow_html=True)

    st.markdown("</div>", unsafe_allow_html=True)
    st.button("🖨️ Imprimer cette page (A4)", key="print_sv",
              on_click=lambda: None,
              help="Utilisez Ctrl+P puis choisissez Format A4")


# ══════════════════════════════════════════════
# ONGLET 2 — SYNTHÈSE SIMPLIFIÉE
# ══════════════════════════════════════════════
with tab2:
    st.markdown('<div class="print-page">', unsafe_allow_html=True)
    st.markdown("""
    <div class="section-header">
      📋 PROJECTION SIMPLIFIÉE — Dispositif Jeanbrun · Compte en T mensuel · Document non contractuel
    </div>
    """, unsafe_allow_html=True)

    # Situation du foyer + opération
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**Situation du foyer**")
        info_foyer = pd.DataFrame({
            "Paramètre": ["Revenus déclarés", "TMI", "Mensualité crédit", "Apport"],
            "Valeur": [
                fmt_eur(v(simp, 'C5')),
                fmt_pct(v(simp, 'C6')),
                fmt_eur(v(simp, 'C7')),
                fmt_eur(v(simp, 'E7')),
            ]
        })
        st.dataframe(info_foyer, hide_index=True, use_container_width=True)

    with c2:
        st.markdown("**Opération immobilière**")
        info_op = pd.DataFrame({
            "Paramètre": ["Prix d'acquisition", "Surface pondérée", "Loyer mensuel initial", "Économie fiscale an 1"],
            "Valeur": [
                fmt_eur(v(simp, 'H5')),
                f"{fmt_num(v(simp, 'H6'), 1)} m²",
                fmt_eur(v(simp, 'H7')),
                fmt_eur(v(simp, 'E6')),
            ]
        })
        st.dataframe(info_op, hide_index=True, use_container_width=True)

    # Tableaux par horizon
    for label, row_base, gain_row, capital_row in [
        ("🔹 Horizon 9 ans", (14, 15, 16, 17), 'M15', 'M14'),
        ("🔸 Horizon 15 ans", (22, 23, 24, 25), 'M23', 'M22'),
        ("⭐ Horizon 25 ans", (30, 31, 32, 33), 'M31', 'M30'),
    ]:
        st.markdown(f"### {label}")
        col_a, col_b, col_c = st.columns([2, 2, 1])
        r0, r1, r2, r3 = row_base
        with col_a:
            st.markdown("**✚ Ce qui rentre**")
            df_in = pd.DataFrame({
                "": ["Loyer mensuel moyen", "Gain fiscal / mois", "Total entrées"],
                "€/mois": [
                    fmt_eur(v(simp, f'D{r0}')),
                    fmt_eur(v(simp, f'D{r1}')),
                    fmt_eur(v(simp, f'D{r2}')),
                ]
            })
            st.dataframe(df_in, hide_index=True, use_container_width=True)
        with col_b:
            st.markdown("**— Ce qui sort**")
            df_out = pd.DataFrame({
                "": ["Mensualité de crédit", "Charges exploitation", "Total sorties"],
                "€/mois": [
                    fmt_eur(v(simp, f'I{r0}')),
                    fmt_eur(v(simp, f'I{r1}')),
                    fmt_eur(v(simp, f'I{r2}')),
                ]
            })
            st.dataframe(df_out, hide_index=True, use_container_width=True)
        with col_c:
            effort = v(simp, f'G{r3}')
            ef = float(effort) if effort else 0
            color = "#EA653D" if ef < 0 else "#009FA3"
            st.markdown(f"""
            <div style="text-align:center; padding:1rem; background:#F4F6F9; border-radius:8px;">
              <div style="font-size:.8rem; color:#555;">Effort mensuel</div>
              <div style="font-size:1.4rem; font-weight:700; color:{color};">{fmt_eur(abs(ef))}</div>
              <hr style="margin:.5rem 0;">
              <div style="font-size:.8rem; color:#555;">Capital constitué</div>
              <div style="font-size:1rem; font-weight:600; color:#13415B;">{fmt_eur(v(simp, capital_row))}</div>
              <div style="font-size:.8rem; color:#555;">Gain fiscal total</div>
              <div style="font-size:1rem; font-weight:600; color:#3761AD;">{fmt_eur(v(simp, gain_row))}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("</div>", unsafe_allow_html=True)
    st.button("🖨️ Imprimer (A4)", key="print_simp")


# ══════════════════════════════════════════════
# ONGLET 3 — SYNTHÈSE DÉTAILLÉE
# ══════════════════════════════════════════════
with tab3:
    st.markdown('<div class="print-page">', unsafe_allow_html=True)
    st.markdown("""
    <div class="section-header">
      📈 PROJECTION FINANCIÈRE ANNUELLE — Dispositif Jeanbrun · Document non contractuel
    </div>
    """, unsafe_allow_html=True)

    # En-tête infos
    c1, c2, c3 = st.columns(3)
    with c1:
        st.metric("Revenus déclarés", fmt_eur(v(det, 'C5')))
        st.metric("TMI", fmt_pct(v(det, 'F6')))
    with c2:
        st.metric("Impôt avant opération (an 1)", fmt_eur(v(det, 'C6')))
        st.metric("Impôt après opération (an 1)", fmt_eur(v(det, 'C7')))
    with c3:
        st.metric("Économie fiscale an 1", fmt_eur(v(det, 'F7')), delta="gain")
        st.metric("Loyer mensuel initial", fmt_eur(v(det, 'J7')))

    st.markdown("---")

    # Tableau annuel
    rows = []
    for i, row_num in enumerate(range(16, 41)):  # 25 ans → lignes 16 à 40
        an = v(det, f'A{row_num}')
        if an is None:
            continue
        rows.append({
            "An": int(float(an)),
            "Loyers perçus (€)": fmt_eur(v(det, f'B{row_num}')),
            "Rembours. prêt (€)": fmt_eur(v(det, f'C{row_num}')),
            "Charges exploit. (€)": fmt_eur(v(det, f'D{row_num}')),
            "Amort. Jeanbrun (€)": fmt_eur(v(det, f'E{row_num}')),
            "RF net imputé (€)": fmt_eur(v(det, f'F{row_num}')),
            "Impôt avant (€)": fmt_eur(v(det, f'G{row_num}')),
            "Impôt après (€)": fmt_eur(v(det, f'H{row_num}')),
            "Éco. fiscale (€)": fmt_eur(v(det, f'I{row_num}')),
            "Effort mensuel (€)": fmt_eur(v(det, f'J{row_num}')),
            "Capital net 0% (€)": fmt_eur(v(det, f'K{row_num}')),
            "Capital net 1,5% (€)": fmt_eur(v(det, f'L{row_num}')),
            "Amt restant (€)": fmt_eur(v(det, f'M{row_num}')),
        })

    if rows:
        df_det = pd.DataFrame(rows)
        st.dataframe(df_det, hide_index=True, use_container_width=True,
                     height=600)
    else:
        st.warning("Données annuelles non disponibles — vérifiez le fichier Excel.")

    st.markdown("</div>", unsafe_allow_html=True)
    st.button("🖨️ Imprimer (A4 paysage)", key="print_det")


# ══════════════════════════════════════════════
# ONGLET 4 — REVENTE & PLUS-VALUE
# ══════════════════════════════════════════════
with tab4:
    st.markdown('<div class="print-page">', unsafe_allow_html=True)
    st.markdown("""
    <div class="section-header">
      🏦 SIMULATION DE REVENTE — Plus-value et enrichissement net · Document non contractuel
    </div>
    """, unsafe_allow_html=True)

    cols = st.columns(3)
    horizons = [
        ("🔹 Revente à 9 ans", "C", "9 ans"),
        ("🔸 Revente à 15 ans", "F", "15 ans"),
        ("⭐ Revente à 25 ans", "I", "25 ans"),
    ]

    for col, (label, col_letter, years) in zip(cols, horizons):
        with col:
            st.markdown(f"#### {label}")
            data = {
                "": [
                    "Prix de vente (0% revalo.)",
                    "Prix de vente (+1,5%/an)",
                    "—",
                    "Prix d'acquisition",
                    "+ Forfait frais acq. (7,5%)",
                    "+ Forfait travaux (15%)",
                    "– Amt réintégrés",
                    "= Prix de revient corrigé",
                    "—",
                    "PV brute (0% revalo.)",
                    "PV brute (+1,5%/an)",
                    "—",
                    "Abattement IR (%)",
                    "PV imposable IR (0%)",
                    "PV imposable IR (+1,5%)",
                    "—",
                    "Impôt PV + PS (0%)",
                    "Impôt PV + PS (+1,5%)",
                ],
                "Valeur": [
                    fmt_eur(v(rev, f'{col_letter}6')),
                    fmt_eur(v(rev, f'{col_letter}7')),
                    "",
                    fmt_eur(v(rev, f'{col_letter}10')),
                    fmt_eur(v(rev, f'{col_letter}11')),
                    fmt_eur(v(rev, f'{col_letter}12')),
                    fmt_eur(v(rev, f'{col_letter}13')),
                    fmt_eur(v(rev, f'{col_letter}14')),
                    "",
                    fmt_eur(v(rev, f'{col_letter}16')),
                    fmt_eur(v(rev, f'{col_letter}17')),
                    "",
                    fmt_pct(v(rev, f'{col_letter}20')),
                    fmt_eur(v(rev, f'{col_letter}21')),
                    fmt_eur(v(rev, f'{col_letter}22')) if v(rev, f'{col_letter}22') else "—",
                    "",
                    fmt_eur(v(rev, f'{col_letter}24')) if v(rev, f'{col_letter}24') else "—",
                    fmt_eur(v(rev, f'{col_letter}25')) if v(rev, f'{col_letter}25') else "—",
                ]
            }
            df_rev = pd.DataFrame(data)
            st.dataframe(df_rev, hide_index=True, use_container_width=True)

    st.markdown("</div>", unsafe_allow_html=True)
    st.button("🖨️ Imprimer (A4)", key="print_rev")


# ══════════════════════════════════════════════
# ONGLET 5 — TÉLÉCHARGEMENT
# ══════════════════════════════════════════════
with tab5:
    st.markdown("### ⬇️ Télécharger le fichier Excel recalculé")
    st.info("Téléchargez le fichier Excel complet avec toutes vos hypothèses et les calculs mis à jour.")

    # Reconstruire le fichier Excel en mémoire pour le download
    tmp_dl = Path(tempfile.mktemp(suffix=".xlsx"))
    wb.save(str(tmp_dl))

    with open(tmp_dl, "rb") as f:
        excel_bytes = f.read()

    tmp_dl.unlink(missing_ok=True)

    st.download_button(
        label="📥 Télécharger la simulation Excel",
        data=excel_bytes,
        file_name=f"Simulation_Jeanbrun_{prix_acq//1000}k_Zone{zone}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    st.markdown("---")
    st.markdown("### ℹ️ Informations")
    st.markdown(f"""
    - **Outil** : Simulateur Dispositif Jeanbrun V9
    - **Moteur de calcul** : LibreOffice Calc (recalcul exact des formules Excel)
    - **Document non contractuel** — à titre pédagogique uniquement
    - Toutes les formules fiscales sont celles du fichier Excel de référence
    """)
